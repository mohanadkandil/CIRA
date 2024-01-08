import cv2
from deepface import DeepFace
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import os

# Load the pre-trained emotion detection model
model = DeepFace.build_model("Emotion")

# Define emotion labels
emotion_labels = ['angry', 'disgust', 'fear', 'happy', 'sad', 'surprise', 'neutral']

# Load face cascade classifier
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Start capturing video
cap = cv2.VideoCapture(1)
emotion_counts = {emotion: 0 for emotion in emotion_labels}
start_time = time.time()
update_interval = 60
excel_file = "emotion_data.xlsx"

timer_start_time = time.time()

# Initialize Excel file if not exists
if not os.path.exists(excel_file):
    pd.DataFrame(columns=['Time'] + emotion_labels).to_excel(excel_file, index=False)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    elapsed_time = time.time() - timer_start_time
    if elapsed_time >= 60:  # Reset timer every 60 seconds
        timer_start_time = time.time()
        elapsed_time = 0

    timer_text = f"Timer: {int(elapsed_time)}s"
    cv2.putText(frame, timer_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

    for (x, y, w, h) in faces:
        face_roi = gray_frame[y:y + h, x:x + w]
        resized_face = cv2.resize(face_roi, (48, 48), interpolation=cv2.INTER_AREA)
        normalized_face = resized_face / 255.0
        reshaped_face = normalized_face.reshape(1, 48, 48, 1)

        preds = model.predict(reshaped_face)[0]
        emotion_idx = preds.argmax()
        emotion = emotion_labels[emotion_idx]
        emotion_counts[emotion] += 1

        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
        cv2.putText(frame, emotion, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

    if time.time() - start_time > update_interval:
        df = pd.DataFrame([emotion_counts])
        emotion_counts = {emotion: 0 for emotion in emotion_labels}  # Reset counts

        # Append data to the Excel file
        if pd.io.common.file_exists(excel_file):
            reader = pd.read_excel(excel_file)
            with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, startrow=reader.shape[0] + 1, index=False, header=False)
        else:
            df.to_excel(excel_file, index=False)

        # Update Chart on the same sheet
        wb = load_workbook(excel_file)
        ws = wb.active

        # Clear existing chart
        for obj in ws._charts:
            ws._charts.remove(obj)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Emotion Count"
        chart.y_axis.title = 'Count'
        chart.x_axis.title = 'Emotions'

        data = Reference(ws, min_col=2, min_row=1, max_col=len(emotion_labels) + 1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "J10")

        wb.save(excel_file)
        start_time = time.time()

    cv2.imshow('Real-time Emotion Detection', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()